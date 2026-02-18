#!/usr/bin/env python3
"""
Sync services from SOW_Services_Reference.xlsx into App.jsx.

Usage:
    python scripts/sync-services.py path/to/SOW_Services_Reference.xlsx [--bump patch|minor]

What it does:
    1. Reads Services Master + Trigger Patterns sheets from the spreadsheet
    2. Regenerates the SERVICE_TRIGGERS JS constant
    3. Regenerates the PRICING_GUIDE template string
    4. Splices both into src/App.jsx between SYNC marker comments
    5. Optionally bumps the version number

The spreadsheet is the single source of truth for services, pricing, bundles,
conditions, and trigger patterns. Edit the spreadsheet, run this script, done.
"""

import sys, os, json, re
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── Helpers ──────────────────────────────────────────────────────────────────

def js_str(s):
    """Escape a string for JS single-quoted string."""
    if not s or s == 'None':
        return ''
    return str(s).replace("\\", "\\\\").replace("'", "\\'").replace("\n", " ").strip()

def fmt_num(v):
    if v is None or v == '' or str(v) == 'None':
        return None
    return int(float(v))

def fmt_pct(v):
    if v is None or v == '' or str(v) == 'None':
        return None
    return float(v)

def parse_trigger_lines(text):
    if not text or str(text) == 'None':
        return []
    return [js_str(t.strip()) for t in str(text).split('\n') if t.strip()]

def fmt_js_arr(items):
    if not items:
        return '[]'
    inner = ', '.join(f"'{i}'" for i in items if i)
    return f'[{inner}]'


# ── Read spreadsheet ────────────────────────────────────────────────────────

def read_spreadsheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # Services Master (cols A-M)
    ws = wb['Services Master']
    services = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if not cat or not name:
            continue
        services.append({
            'category': str(cat).strip(),
            'serviceName': str(name).strip(),
            'recommend': str(ws.cell(r, 3).value or 'conditional').strip(),
            'condition': str(ws.cell(r, 4).value or '').strip(),
            'bundle': str(ws.cell(r, 5).value or '').strip(),
            'engagementType': str(ws.cell(r, 6).value or 'fixed_fee').strip(),
            'termLow': ws.cell(r, 7).value,
            'termHigh': ws.cell(r, 8).value,
            'budgetLow': ws.cell(r, 9).value,
            'budgetHigh': ws.cell(r, 10).value,
            'pctProject': ws.cell(r, 11).value,
            'pctPaidMedia': ws.cell(r, 12).value,
            'note': str(ws.cell(r, 13).value or '').strip(),
        })

    # Trigger Patterns
    ws2 = wb['Trigger Patterns']
    triggers = {}
    for r in range(2, ws2.max_row + 1):
        cid = ws2.cell(r, 1).value
        if not cid:
            continue
        cid = str(cid).strip()
        triggers[str(ws2.cell(r, 2).value or '').strip()] = {
            'id': cid,
            'category': str(ws2.cell(r, 2).value or '').strip(),
            'description': str(ws2.cell(r, 3).value or '').strip(),
            'engagementType': str(ws2.cell(r, 4).value or '').strip(),
            'direct': parse_trigger_lines(ws2.cell(r, 5).value),
            'indirect': parse_trigger_lines(ws2.cell(r, 6).value),
            'situational': parse_trigger_lines(ws2.cell(r, 7).value),
            'performance': parse_trigger_lines(ws2.cell(r, 8).value),
            'sampleLanguage': parse_trigger_lines(ws2.cell(r, 9).value),
        }

    return services, triggers


# ── Generate SERVICE_TRIGGERS JS ────────────────────────────────────────────

def generate_service_triggers(services, triggers):
    # Group by category, preserving order
    cat_services = OrderedDict()
    for s in services:
        cat = s['category']
        if cat not in cat_services:
            cat_services[cat] = []
        cat_services[cat].append(s)

    lines = []
    lines.append("// === SYNC:SERVICE_TRIGGERS_START ===")
    lines.append("const SERVICE_TRIGGERS = [")

    for cat_name, svcs in cat_services.items():
        eng_type = svcs[0]['engagementType']
        tp = triggers.get(cat_name)

        # Derive ID from trigger patterns or generate from name
        if tp:
            cat_id = tp['id']
            desc = js_str(tp.get('description', cat_name))
        else:
            cat_id = cat_name.lower().replace(' & ', '_').replace(' ', '_').replace('(', '').replace(')', '')
            desc = js_str(cat_name)

        lines.append(f"  {{")
        lines.append(f"    id: '{cat_id}',")
        lines.append(f"    category: '{js_str(cat_name)}',")
        lines.append(f"    description: '{desc}',")
        lines.append(f"    engagementType: '{eng_type}',")
        lines.append(f"    services: [")

        current_bundle = None
        for s in svcs:
            name = js_str(s['serviceName'])
            rec = s['recommend'] if s['recommend'] not in ('', 'None') else 'conditional'
            cond = js_str(s['condition'])
            bundle = s['bundle'] if s['bundle'] not in ('', 'None') else ''
            note = s['note'] if s['note'] not in ('', 'None') else ''

            if bundle and bundle != current_bundle:
                current_bundle = bundle
                lines.append(f"      // {bundle} bundle")
            elif not bundle and current_bundle:
                current_bundle = None
                lines.append(f"      // Individual services")

            pricing_parts = []
            tl = fmt_num(s['termLow']); th = fmt_num(s['termHigh'])
            bl = fmt_num(s['budgetLow']); bh = fmt_num(s['budgetHigh'])
            pp = fmt_pct(s['pctProject']); pm = fmt_pct(s['pctPaidMedia'])

            if tl is not None: pricing_parts.append(f"termLow: {tl}")
            if th is not None: pricing_parts.append(f"termHigh: {th}")
            if bl is not None: pricing_parts.append(f"budgetLow: {bl}")
            if bh is not None: pricing_parts.append(f"budgetHigh: {bh}")
            if pp is not None: pricing_parts.append(f"percentageOfProject: {int(pp) if pp == int(pp) else pp}")
            if pm is not None: pricing_parts.append(f"percentageOfPaidMedia: {int(pm) if pm == int(pm) else pm}")
            if bundle: pricing_parts.append(f"bundle: '{js_str(bundle)}'")
            if note: pricing_parts.append(f"note: '{js_str(note)}'")

            pricing_str = '{ ' + ', '.join(pricing_parts) + ' }' if pricing_parts else '{}'
            lines.append(f"      {{ name: '{name}', recommend: '{rec}', condition: '{cond}', pricing: {pricing_str} }},")

        lines.append(f"    ],")

        # Trigger patterns
        if tp:
            tp_data = tp
        else:
            tp_data = {'direct': [], 'indirect': [], 'situational': [], 'performance': [], 'sampleLanguage': []}

        lines.append(f"    triggerPatterns: {{")
        lines.append(f"      direct: {fmt_js_arr(tp_data['direct'])},")
        lines.append(f"      indirect: {fmt_js_arr(tp_data['indirect'])},")
        lines.append(f"      situational: {fmt_js_arr(tp_data['situational'])},")
        lines.append(f"      performance: {fmt_js_arr(tp_data['performance'])},")
        lines.append(f"      sampleLanguage: {fmt_js_arr(tp_data['sampleLanguage'])}")
        lines.append(f"    }}")
        lines.append(f"  }},")

    lines.append("];")
    lines.append("// === SYNC:SERVICE_TRIGGERS_END ===")
    return '\n'.join(lines)


# ── Generate PRICING_GUIDE ──────────────────────────────────────────────────

def generate_pricing_guide(services, triggers):
    ENG_LABELS = {'fixed_fee': 'Fixed Fee', 'retainer': 'Retainer', 'tm': 'Time & Materials', 'any': 'Any'}

    cats = OrderedDict()
    for s in services:
        cat = s['category']
        if cat not in cats:
            cats[cat] = {'eng': s['engagementType'], 'services': []}
        cats[cat]['services'].append(s)

    lines = []
    lines.append("// === SYNC:PRICING_GUIDE_START ===")
    lines.append("const PRICING_GUIDE = `")
    lines.append("## SERVICE PRICING GUIDE - For Validation")
    lines.append("")
    lines.append("Use this guide to validate pricing in SOWs. Flag fees that are significantly below the low range (underpriced) or above the high range (overpriced).")

    for cat_name, cat_data in cats.items():
        eng = ENG_LABELS.get(cat_data['eng'], cat_data['eng'])
        lines.append("")
        lines.append(f"### {cat_name} ({eng})")
        lines.append("| Service | Term | Budget Range |")
        lines.append("|---------|------|--------------|")

        for s in cat_data['services']:
            bundle = s['bundle'] if s['bundle'] not in ('', 'None') else ''
            note = s['note'] if s['note'] not in ('', 'None') else ''
            bl = s['budgetLow']; bh = s['budgetHigh']
            tl = s['termLow']; th = s['termHigh']
            pp = s['pctProject']; pm = s['pctPaidMedia']

            # Skip sub-bundle items with no pricing
            if bundle and not bl and not pp and not pm:
                continue

            # Format term
            if tl and th:
                tli, thi = int(float(tl)), int(float(th))
                if tli == 52 and thi == 52:
                    term = "Annual"
                elif tli == thi:
                    term = f"{tli} weeks"
                else:
                    term = f"{tli}-{thi} weeks"
            else:
                term = "Varies"

            # Format budget
            if pp:
                budget = f"~{int(float(pp))}% of total project fee"
            elif pm:
                budget = f"~{int(float(pm))}% of paid media management fees"
            elif bl and bh:
                budget = f"${int(float(bl)):,} - ${int(float(bh)):,}"
                if note: budget += f" ({note})"
            else:
                budget = "T&M based on scope"

            display_name = s['serviceName']
            if bundle:
                display_name = f"{display_name} [{bundle}]"

            lines.append(f"| {display_name} | {term} | {budget} |")

    lines.append("`;")
    lines.append("// === SYNC:PRICING_GUIDE_END ===")
    return '\n'.join(lines)


# ── Splice into App.jsx ─────────────────────────────────────────────────────

def splice_between_markers(source, start_marker, end_marker, replacement):
    """Replace everything between start_marker and end_marker (inclusive) with replacement."""
    pattern = re.compile(
        re.escape(start_marker) + r'.*?' + re.escape(end_marker),
        re.DOTALL
    )
    result, count = pattern.subn(replacement, source)
    if count == 0:
        print(f"  ERROR: Markers not found: {start_marker} ... {end_marker}")
        sys.exit(1)
    return result


def bump_version(source, bump_type='patch'):
    """Bump APP_VERSION in source."""
    match = re.search(r"const APP_VERSION = '(\d+)\.(\d+)\.(\d+)';", source)
    if not match:
        print("  WARNING: Could not find APP_VERSION to bump")
        return source
    major, minor, patch = int(match.group(1)), int(match.group(2)), int(match.group(3))
    if bump_type == 'minor':
        minor += 1; patch = 0
    else:
        patch += 1
    new_ver = f"{major}.{minor}.{patch}"
    result = source.replace(match.group(0), f"const APP_VERSION = '{new_ver}';")
    print(f"  Version: {match.group(1)}.{match.group(2)}.{match.group(3)} → {new_ver}")
    return result


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/sync-services.py <spreadsheet.xlsx> [--bump patch|minor]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    bump_type = 'patch'
    if '--bump' in sys.argv:
        idx = sys.argv.index('--bump')
        if idx + 1 < len(sys.argv):
            bump_type = sys.argv[idx + 1]

    # Find App.jsx
    script_dir = os.path.dirname(os.path.abspath(__file__))
    app_path = os.path.join(script_dir, '..', 'src', 'App.jsx')
    app_path = os.path.normpath(app_path)

    if not os.path.exists(xlsx_path):
        print(f"ERROR: Spreadsheet not found: {xlsx_path}")
        sys.exit(1)
    if not os.path.exists(app_path):
        print(f"ERROR: App.jsx not found: {app_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    services, triggers = read_spreadsheet(xlsx_path)
    print(f"  Found {len(services)} services, {len(triggers)} trigger sets")

    print("Generating SERVICE_TRIGGERS...")
    triggers_js = generate_service_triggers(services, triggers)

    print("Generating PRICING_GUIDE...")
    pricing_js = generate_pricing_guide(services, triggers)

    print(f"Updating: {app_path}")
    source = open(app_path, 'r').read()

    source = splice_between_markers(
        source,
        '// === SYNC:SERVICE_TRIGGERS_START ===',
        '// === SYNC:SERVICE_TRIGGERS_END ===',
        triggers_js
    )
    print("  ✓ SERVICE_TRIGGERS replaced")

    source = splice_between_markers(
        source,
        '// === SYNC:PRICING_GUIDE_START ===',
        '// === SYNC:PRICING_GUIDE_END ===',
        pricing_js
    )
    print("  ✓ PRICING_GUIDE replaced")

    source = bump_version(source, bump_type)

    open(app_path, 'w').write(source)

    # Summary
    from collections import Counter
    cats = Counter(s['category'] for s in services)
    print(f"\nSync complete:")
    print(f"  {len(services)} services across {len(cats)} categories")
    print(f"  {len(triggers)} trigger pattern sets")
    print(f"  App.jsx updated at {app_path}")
    print(f"\nNext: npm run build (or npx vite build) to verify")


if __name__ == '__main__':
    main()
